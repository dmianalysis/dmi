#!/usr/bin/env python3
"""Import the official BLS UCC -> ELI concordance into a pinned, normalized
research artifact.

Detailed Inflation Substrate v0.1, Milestone 1 (prompt section 3).

Why an importer rather than parsing at audit time
-------------------------------------------------
The BLS publishes the concordance as an appendix workbook (and historically as
an HTML appendix). Both formats carry presentation rows -- a merged title row,
a footnote row, a "Note:" row and a source attribution row -- that are brittle
to parse repeatedly. This script converts the workbook once into a plain TSV
that the audit joins against by exact identifier, and records provenance
(source filename, sheet, byte size, SHA-256, row counts, BLS note text) in a
sidecar JSON so the pinned artifact can always be traced back to its origin.

The normalized artifact is small (a few hundred rows) and IS committed. The
source workbook and the CE microdata extracts are external research inputs and
are NOT committed.

Usage
-----
    python scripts/import_ucc_eli_concordance.py \
        --source /path/to/ce-cpi-concordance-August-2026.xlsx \
        --out registry/research/ucc_eli_concordance_2024_v0_1.tsv

Attribution: the concordance is a publication of the U.S. Bureau of Labor
Statistics. DMI did not originate it.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path

UCC_RE = re.compile(r"^[0-9]{6}$")
ELI_RE = re.compile(r"^[A-Z]{2}[0-9]{3}$")

EXPECTED_HEADER = ("UCC", "ELI", "UCC Title", "ELI Title", "CE SOURCE")

# BLS marks UCCs that are allocated across several ELIs with a trailing "(1)"
# footnote reference inside the title cell. The marker is presentation, not
# data; it is stripped and recorded as a boolean instead.
FOOTNOTE_MARKER = "(1)"


@dataclass(frozen=True)
class ConcordanceRow:
    ucc: str
    eli: str
    ucc_title: str
    eli_title: str
    ce_source: str
    multi_eli_footnote: bool


def _clean(cell: object) -> str:
    """Normalize a workbook cell to a stripped string.

    BLS cells contain non-breaking spaces (U+00A0); treat them as whitespace.
    """
    if cell is None:
        return ""
    return str(cell).replace("\u00a0", " ").strip()


def parse_workbook(source: Path) -> tuple[list[ConcordanceRow], list[str], str]:
    """Parse the concordance workbook.

    Returns ``(rows, trailer_lines, sheet_name)``. ``trailer_lines`` are the
    BLS footnote/note/source lines that follow the data block; they are
    preserved verbatim in the provenance sidecar.
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - environment guard
        raise SystemExit(
            "openpyxl is required to import the concordance workbook: "
            f"{exc}"
        )

    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    if len(workbook.sheetnames) != 1:
        raise SystemExit(
            f"Expected exactly one sheet in {source.name}, found "
            f"{workbook.sheetnames!r}. Refusing to guess which sheet is "
            "authoritative."
        )
    sheet_name = workbook.sheetnames[0]
    sheet = workbook[sheet_name]

    raw = [tuple(_clean(c) for c in row) for row in sheet.iter_rows(values_only=True)]

    header_index = _find_header(raw, source)
    rows: list[ConcordanceRow] = []
    trailer: list[str] = []

    for line in raw[header_index + 1:]:
        ucc, eli = line[0], line[1]
        if not ucc and not eli:
            continue
        if not eli:
            # A populated first column with no ELI is a BLS trailer line
            # (footnote, note, or source attribution).
            trailer.append(ucc)
            continue
        if not UCC_RE.match(ucc):
            raise SystemExit(
                f"{source.name}: unparseable UCC {ucc!r} on a row that also "
                f"carries ELI {eli!r}. Refusing to import a malformed "
                "concordance."
            )
        if not ELI_RE.match(eli):
            raise SystemExit(
                f"{source.name}: unparseable ELI {eli!r} for UCC {ucc!r}. "
                "Refusing to import a malformed concordance."
            )
        ucc_title = line[2]
        footnoted = ucc_title.endswith(FOOTNOTE_MARKER)
        if footnoted:
            ucc_title = ucc_title[: -len(FOOTNOTE_MARKER)].strip()
        rows.append(
            ConcordanceRow(
                ucc=ucc,
                eli=eli,
                ucc_title=ucc_title,
                eli_title=line[3],
                ce_source=line[4],
                multi_eli_footnote=footnoted,
            )
        )

    if not rows:
        raise SystemExit(f"{source.name}: no concordance rows parsed.")
    return rows, trailer, sheet_name


def _find_header(raw: list[tuple[str, ...]], source: Path) -> int:
    for index, line in enumerate(raw):
        if tuple(line[: len(EXPECTED_HEADER)]) == EXPECTED_HEADER:
            return index
    raise SystemExit(
        f"{source.name}: could not locate the header row "
        f"{EXPECTED_HEADER!r}. The BLS layout may have changed; update this "
        "importer deliberately rather than loosening the match."
    )


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_tsv(rows: list[ConcordanceRow], out: Path) -> None:
    out.parent.mkdir(parents=True, exist_ok=True)
    ordered = sorted(rows, key=lambda r: (r.ucc, r.eli))
    with out.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle, delimiter="\t", lineterminator="\n")
        writer.writerow(
            ["ucc", "eli", "ucc_title", "eli_title", "ce_source", "multi_eli_footnote"]
        )
        for row in ordered:
            writer.writerow(
                [
                    row.ucc,
                    row.eli,
                    row.ucc_title,
                    row.eli_title,
                    row.ce_source,
                    "true" if row.multi_eli_footnote else "false",
                ]
            )


def build_provenance(
    source: Path,
    out: Path,
    rows: list[ConcordanceRow],
    trailer: list[str],
    sheet_name: str,
) -> dict:
    multi = {}
    for row in rows:
        multi.setdefault(row.ucc, []).append(row.eli)
    return {
        "artifact_id": "ucc_eli_concordance_2024_v0_1",
        "version": "0.1.0",
        "status": "RESEARCH_ONLY",
        "normalized_artifact": out.name,
        "publisher": "U.S. Bureau of Labor Statistics",
        "publication": (
            "Appendix 5. Consumer Expenditure survey item name (universal "
            "classification codes-UCC) to Consumer Price Index item titles "
            "(entry level item-ELI) concordance"
        ),
        "source_file_name": source.name,
        "source_sheet": sheet_name,
        "source_sha256": sha256(source),
        "source_bytes": source.stat().st_size,
        "row_count": len(rows),
        "distinct_ucc_count": len(multi),
        "distinct_eli_count": len({r.eli for r in rows}),
        "multi_eli_ucc_count": sum(1 for v in multi.values() if len(v) > 1),
        "bls_trailer_lines": trailer,
        "pinning_note": (
            "This is the August-2026 concordance, aligned to the CPI item "
            "structure introduced for 2024 annual expenditure weights used in "
            "indexes starting January 2026. It must not be substituted with "
            "the archived January-2025 or January-2024 concordances, which "
            "describe a different CPI item structure."
        ),
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument(
        "--source", type=Path, required=True,
        help="Path to the BLS concordance workbook (.xlsx).",
    )
    parser.add_argument(
        "--out", type=Path,
        default=Path("registry/research/ucc_eli_concordance_2024_v0_1.tsv"),
        help="Destination for the normalized TSV artifact.",
    )
    args = parser.parse_args(argv)

    if not args.source.is_file():
        raise SystemExit(f"Concordance source not found: {args.source}")

    rows, trailer, sheet_name = parse_workbook(args.source)
    write_tsv(rows, args.out)

    provenance = build_provenance(args.source, args.out, rows, trailer, sheet_name)
    prov_path = args.out.with_suffix(".provenance.json")
    prov_path.write_text(
        json.dumps(provenance, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )

    print(f"Wrote {args.out} ({provenance['row_count']} rows)")
    print(f"Wrote {prov_path}")
    print(
        f"  distinct UCCs: {provenance['distinct_ucc_count']}  "
        f"distinct ELIs: {provenance['distinct_eli_count']}  "
        f"multi-ELI UCCs: {provenance['multi_eli_ucc_count']}"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
