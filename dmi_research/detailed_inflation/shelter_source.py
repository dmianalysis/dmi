"""Re-derive the shelter-UCC source observations from the pinned 2024 PUMD.

Detailed Inflation Substrate v0.1, shelter task, Phase C1.

RESEARCH ONLY. Nothing here is imported by ``dmi_calculator``, by the
operational Baseline or Slack-Plus paths, or by any release workflow.

Milestone 2 graded ``pumd_membership = VERIFIED`` for UCCs 910104-910107 on
the strength of a manual scan performed in an earlier working session and
transcribed into ``registry/research/ucc_provenance_classes_v0_1.json``. That
record was explicit about its own weakness: ``reproduced_by_test: false``, and
"until that is done and committed, this stays a named prior finding rather
than a check anything re-runs".

This module is that check. It re-reads the same five MTBI members of the same
pinned archive, verifies each member's SHA-256 against the source registry
before believing a single row of it, and re-derives the record counts and
PUBFLAG values independently of the transcribed numbers. The transcribed
numbers are then compared to what was measured, rather than being assumed.

Three disciplines are load-bearing.

*The archive is verified, not merely located.* A count taken from an unpinned
file would reproduce nothing. :func:`verify_archive_members` hashes each MTBI
member against ``pumd_2024_interview_source_v0_1.json`` and refuses to
proceed on a mismatch.

*Counting is separated from judging.* :func:`read_shelter_source` measures and
records; :func:`compare_with_prior_observation` decides whether a measurement
agrees with the preserved claim. Neither function can quietly become the
other.

*Two different record counts exist and are both reported.* The preserved
observation counted every row bearing the UCC, across all reference years. The
confirmed estimator consumes only rows with ``REF_YR == 2024``. Those are
different populations and they differ substantially. Reporting only the one
that matches the preserved claim would reproduce a number while hiding the
number that actually feeds an estimate, so both are carried on every
observation and the estimator-relevant one is named as such.

This module computes no expenditure. It reads COST for no purpose and does not
sum it. Membership and publication status are what C1 is for; amounts are
Phase C3/C4's business and are produced by the frozen estimator, not here.
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Mapping, Sequence

from . import pumd

REPO_ROOT = Path(__file__).resolve().parents[2]

#: The source registry that pins the archive and its members.
SOURCE_REGISTRY_PATH = (
    REPO_ROOT / "registry/research/pumd_2024_interview_source_v0_1.json"
)

#: The Milestone-2 registry carrying the preserved manual observation.
PROVENANCE_REGISTRY_PATH = (
    REPO_ROOT / "registry/research/ucc_provenance_classes_v0_1.json"
)

#: The four CONCORDANCE_ONLY rental-equivalence UCCs. These are the Track-A
#: inputs, in the sense that they are the codes the CPI concordance names.
SHELTER_UCCS: tuple[str, ...] = ("910104", "910105", "910106", "910107")

#: Their published cx.item counterparts. Recorded for contrast only. No
#: Track-A amount may be taken from these, and no estimate for a shelter UCC
#: may be inferred from its counterpart. They are read here because the
#: PUBFLAG contrast is the whole point of the corroboration: it is only
#: evidence if the other side of the partition carries the other value.
PUBLISHED_COUNTERPART_UCCS: tuple[str, ...] = ("910050", "910101", "910102", "910103")

OBSERVED_UCCS: tuple[str, ...] = SHELTER_UCCS + PUBLISHED_COUNTERPART_UCCS

#: The preserved prior manual observation, transcribed verbatim from
#: ``ucc_provenance_classes_v0_1.json :: pumd_observations
#: .CE_2024_INTERVIEW_MTBI_SHELTER_RENTAL_EQUIVALENCE``. These are the claims
#: under test. They are stated here as data so the comparison is against a
#: fixed target and cannot be adjusted to whatever was measured.
PRIOR_MANUAL_RECORD_COUNTS: Mapping[str, int] = {
    "910104": 46119,
    "910105": 2070,
    "910106": 45,
    "910107": 891,
}

#: The preserved PUBFLAG claim: uniform within each UCC, 1 for the four
#: concordance-only codes and 2 for the four published counterparts.
PRIOR_MANUAL_PUBFLAG: Mapping[str, str] = {
    "910104": "1",
    "910105": "1",
    "910106": "1",
    "910107": "1",
    "910050": "2",
    "910101": "2",
    "910102": "2",
    "910103": "2",
}

#: What the prior observation counted. Recorded because the answer turns out
#: to matter: see :attr:`ShelterSourceObservation.records_in_benchmark_year`.
PRIOR_MANUAL_COUNTING_BASIS = "ALL_REFERENCE_YEARS"

#: PUBFLAG code meanings, as published in the BLS PUMD data dictionary. Stated
#: here as the expected value; :func:`read_pubflag_dictionary` re-reads them
#: from the workbook rather than trusting this constant.
EXPECTED_PUBFLAG_MEANINGS: Mapping[str, str] = {
    "1": "Not published",
    "2": "Published in Integrated Bulletin",
}

#: Where the dictionary workbook lives, relative to the PUMD data root. The
#: workbook is not redistributable and is not committed.
PUBFLAG_DICTIONARY_RELATIVE_PATH = "docs/ce-pumd-interview-diary-dictionary.xlsx"

#: The sheet and columns the code meanings are read from.
PUBFLAG_DICTIONARY_SHEET = "Codes "
PUBFLAG_DICTIONARY_SURVEY = "INTERVIEW"
PUBFLAG_DICTIONARY_FILE = "MTBI"
PUBFLAG_DICTIONARY_VARIABLE = "PUBFLAG"

#: Reproduction verdicts.
MATCH = "MATCH"
DIFFERS = "DIFFERS"


class ShelterSourceError(RuntimeError):
    """The pinned source does not support a check this module must make.

    Raised, never swallowed. C1 exists to stop the milestone if the archive
    has moved under the preserved observation; absorbing that would defeat
    the entire purpose of the phase.
    """


# --------------------------------------------------------------------------
# Stage 1: verify the archive is the pinned one
# --------------------------------------------------------------------------


def file_digest(path: str | os.PathLike[str]) -> str:
    """SHA-256 of a file, read in blocks so a 50 MB CSV is not held in RAM."""
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for block in iter(lambda: handle.read(1 << 20), b""):
            digest.update(block)
    return digest.hexdigest()


def pinned_member_digests() -> dict[str, str]:
    """The MTBI member SHA-256 values recorded in the source registry."""
    registry = json.loads(SOURCE_REGISTRY_PATH.read_text(encoding="utf-8"))
    members = registry["archives"]["INTRVW24"]["members_relied_on"]
    return {
        Path(name).name: digest
        for name, digest in members.items()
        if Path(name).name.startswith("mtbi")
    }


def verify_archive_members(directory: str | os.PathLike[str]) -> dict[str, str]:
    """Hash every MTBI member and compare against the pinned registry.

    Returns the measured digests. Raises on any mismatch, on any missing
    file, and on any registry-pinned member the estimator's file list does
    not name - the last of those catching a registry and a file list that
    have drifted apart, which would otherwise silently narrow the scan.
    """
    base = Path(directory)
    pinned = pinned_member_digests()
    expected_names = {name for _fmli, name, _role in pumd.QUARTER_FILES}

    missing_from_registry = sorted(expected_names - pinned.keys())
    if missing_from_registry:
        raise ShelterSourceError(
            "the source registry does not pin these MTBI members that the "
            f"estimator reads: {missing_from_registry}"
        )
    missing_from_estimator = sorted(pinned.keys() - expected_names)
    if missing_from_estimator:
        raise ShelterSourceError(
            "the source registry pins MTBI members the estimator does not "
            f"read: {missing_from_estimator}"
        )

    measured: dict[str, str] = {}
    mismatched: list[str] = []
    for name in sorted(expected_names):
        path = base / name
        if not path.exists():
            raise pumd.PumdDataUnavailable(f"MTBI member not present: {path}")
        digest = file_digest(path)
        measured[name] = digest
        if digest != pinned[name]:
            mismatched.append(f"{name}: pinned {pinned[name]}, measured {digest}")
    if mismatched:
        raise ShelterSourceError(
            "the local MTBI files are not the pinned archive members; a count "
            "taken from them would reproduce nothing. Differences: "
            + "; ".join(mismatched)
        )
    return measured


# --------------------------------------------------------------------------
# Stage 2: count
# --------------------------------------------------------------------------


@dataclass(frozen=True)
class ShelterSourceObservation:
    """What the pinned MTBI files actually say about one UCC.

    Every field is a count or a tally of codes. No expenditure is recorded,
    because an unweighted COST sum is not an estimate and recording one here
    would invite it being read as one.
    """

    ucc: str

    #: Rows bearing this UCC across all five members, regardless of REF_YR.
    #: This is the basis the preserved manual observation used.
    records_all_reference_years: int

    #: Rows bearing this UCC with ``REF_YR == 2024``. This is the basis the
    #: confirmed estimator uses, because the BLS sample program filters on
    #: reference year. It is the count that bears on any figure produced in
    #: Phase C4, and it is smaller than the count above for every UCC here.
    records_in_benchmark_year: int

    distinct_newids_all_reference_years: int
    distinct_newids_in_benchmark_year: int

    #: PUBFLAG value -> row count, over all reference years.
    pubflag_tally: Mapping[str, int]

    #: MTBI member file name -> row count, over all reference years.
    records_by_file: Mapping[str, int]

    #: REF_YR -> row count.
    records_by_reference_year: Mapping[int, int]

    @property
    def pubflag_is_uniform(self) -> bool:
        return len(self.pubflag_tally) == 1

    @property
    def sole_pubflag(self) -> str | None:
        if not self.pubflag_is_uniform:
            return None
        return next(iter(self.pubflag_tally))


def read_shelter_source(
    directory: str | os.PathLike[str],
    uccs: Sequence[str] = OBSERVED_UCCS,
    year: int = pumd.BENCHMARK_YEAR,
) -> dict[str, ShelterSourceObservation]:
    """Scan the five MTBI members once and tally the requested UCCs.

    This does not reuse :func:`pumd.read_mtbi`, for two reasons that are both
    about not weakening the check. That function drops PUBFLAG, which is half
    of what C1 must verify, and it applies the calendar-year filter, which
    would make the all-reference-year count unmeasurable. The file list,
    the year predicate and the UCC normalisation are all taken from
    :mod:`pumd` so that no part of the reading contract is re-decided here.
    """
    base = Path(directory)
    wanted = frozenset(uccs)
    required = ("NEWID", "UCC", "REF_YR", "REF_MO", "PUBFLAG")

    all_rows: Counter[str] = Counter()
    year_rows: Counter[str] = Counter()
    pubflags: dict[str, Counter[str]] = {u: Counter() for u in wanted}
    by_file: dict[str, Counter[str]] = {u: Counter() for u in wanted}
    by_year: dict[str, Counter[int]] = {u: Counter() for u in wanted}
    newids_all: dict[str, set[str]] = {u: set() for u in wanted}
    newids_year: dict[str, set[str]] = {u: set() for u in wanted}

    for _fmli_name, mtbi_name, _role in pumd.QUARTER_FILES:
        path = base / mtbi_name
        if not path.exists():
            raise pumd.PumdDataUnavailable(f"MTBI member not present: {path}")
        with open(path, newline="", encoding="utf-8", errors="replace") as handle:
            reader = csv.DictReader(handle)
            fieldnames = reader.fieldnames or []
            absent = [c for c in required if c not in fieldnames]
            if absent:
                raise pumd.PumdSchemaError(
                    f"{mtbi_name} is missing columns: {', '.join(absent)}"
                )
            for row in reader:
                ucc = row["UCC"].strip()
                if ucc not in wanted:
                    continue
                reference_year = int(row["REF_YR"])
                newid = row["NEWID"].strip()
                all_rows[ucc] += 1
                pubflags[ucc][row["PUBFLAG"].strip()] += 1
                by_file[ucc][mtbi_name] += 1
                by_year[ucc][reference_year] += 1
                newids_all[ucc].add(newid)
                if pumd.is_in_benchmark_year(reference_year, year):
                    year_rows[ucc] += 1
                    newids_year[ucc].add(newid)

    return {
        ucc: ShelterSourceObservation(
            ucc=ucc,
            records_all_reference_years=all_rows[ucc],
            records_in_benchmark_year=year_rows[ucc],
            distinct_newids_all_reference_years=len(newids_all[ucc]),
            distinct_newids_in_benchmark_year=len(newids_year[ucc]),
            pubflag_tally=dict(sorted(pubflags[ucc].items())),
            records_by_file=dict(by_file[ucc]),
            records_by_reference_year=dict(sorted(by_year[ucc].items())),
        )
        for ucc in sorted(wanted)
    }


# --------------------------------------------------------------------------
# Stage 3: judge
# --------------------------------------------------------------------------


@dataclass(frozen=True)
class ReproductionCheck:
    """One preserved claim, what was measured against it, and the verdict."""

    claim: str
    ucc: str
    claimed: object
    measured: object
    status: str
    note: str = ""

    @property
    def reproduced(self) -> bool:
        return self.status == MATCH


def compare_with_prior_observation(
    observations: Mapping[str, ShelterSourceObservation],
) -> list[ReproductionCheck]:
    """Compare every preserved claim against what was measured.

    The record-count claims are compared on the all-reference-year basis,
    because that is the basis the preserved observation used. Comparing them
    against the calendar-year count would manufacture a discrepancy out of a
    definitional difference; comparing them against whichever basis agreed
    would be worse still. The basis is fixed in advance by
    :data:`PRIOR_MANUAL_COUNTING_BASIS` and stated on every check.
    """
    checks: list[ReproductionCheck] = []

    for ucc in SHELTER_UCCS:
        observation = observations.get(ucc)
        if observation is None:
            checks.append(
                ReproductionCheck(
                    claim="RECORD_COUNT",
                    ucc=ucc,
                    claimed=PRIOR_MANUAL_RECORD_COUNTS[ucc],
                    measured=None,
                    status=DIFFERS,
                    note="the UCC was not scanned",
                )
            )
            continue
        measured = observation.records_all_reference_years
        claimed = PRIOR_MANUAL_RECORD_COUNTS[ucc]
        checks.append(
            ReproductionCheck(
                claim="RECORD_COUNT",
                ucc=ucc,
                claimed=claimed,
                measured=measured,
                status=MATCH if measured == claimed else DIFFERS,
                note=(
                    f"counting basis {PRIOR_MANUAL_COUNTING_BASIS}; the "
                    f"calendar-year-eligible count the estimator consumes is "
                    f"{observation.records_in_benchmark_year}"
                ),
            )
        )

    for ucc, claimed_flag in PRIOR_MANUAL_PUBFLAG.items():
        observation = observations.get(ucc)
        if observation is None:
            checks.append(
                ReproductionCheck(
                    claim="PUBFLAG",
                    ucc=ucc,
                    claimed=claimed_flag,
                    measured=None,
                    status=DIFFERS,
                    note="the UCC was not scanned",
                )
            )
            continue
        if not observation.pubflag_is_uniform:
            checks.append(
                ReproductionCheck(
                    claim="PUBFLAG",
                    ucc=ucc,
                    claimed=claimed_flag,
                    measured=dict(observation.pubflag_tally),
                    status=DIFFERS,
                    note=(
                        "the preserved observation recorded PUBFLAG as uniform "
                        "within this UCC; it is not uniform in the pinned files"
                    ),
                )
            )
            continue
        measured_flag = observation.sole_pubflag
        checks.append(
            ReproductionCheck(
                claim="PUBFLAG",
                ucc=ucc,
                claimed=claimed_flag,
                measured=measured_flag,
                status=MATCH if measured_flag == claimed_flag else DIFFERS,
                note=f"uniform across {observation.records_all_reference_years} rows",
            )
        )

    return checks


def partition_is_corroborated(
    observations: Mapping[str, ShelterSourceObservation],
) -> bool:
    """Does PUBFLAG split exactly on the cx.item membership partition?

    True only when every concordance-only shelter UCC is uniformly PUBFLAG=1
    and every published counterpart is uniformly PUBFLAG=2. This is the claim
    the Milestone-2 registry calls "independent support for the classification
    from a source the classification does not use". It is checked rather than
    restated.
    """
    for ucc in SHELTER_UCCS:
        observation = observations.get(ucc)
        if observation is None or observation.sole_pubflag != "1":
            return False
    for ucc in PUBLISHED_COUNTERPART_UCCS:
        observation = observations.get(ucc)
        if observation is None or observation.sole_pubflag != "2":
            return False
    return True


# --------------------------------------------------------------------------
# Stage 4: the dictionary meaning of PUBFLAG
# --------------------------------------------------------------------------


@dataclass(frozen=True)
class PubflagDictionaryReading:
    """The MTBI PUBFLAG code meanings, read from the BLS data dictionary."""

    workbook_sha256: str
    workbook_sha256_matches_registry: bool
    sheet: str
    variable_description: str
    code_meanings: Mapping[str, str]
    rows_read: tuple[int, ...] = field(default=())

    @property
    def agrees_with_expectation(self) -> bool:
        return dict(self.code_meanings) == dict(EXPECTED_PUBFLAG_MEANINGS)


def pinned_dictionary_digest() -> str:
    registry = json.loads(SOURCE_REGISTRY_PATH.read_text(encoding="utf-8"))
    return registry["documents"]["PUMD_DICTIONARY"]["sha256"]


def read_pubflag_dictionary(path: str | os.PathLike[str]) -> PubflagDictionaryReading:
    """Read the MTBI PUBFLAG code meanings out of the BLS dictionary workbook.

    The meanings are read from the workbook, not asserted from
    :data:`EXPECTED_PUBFLAG_MEANINGS`, which exists only so the reading can be
    contradicted. Rows are filtered to survey INTERVIEW, file MTBI and
    variable PUBFLAG; the dictionary carries the same variable on ITBI, ITII
    and PUBF, and those are different files whose codes are not what the
    expenditure records carry.

    Only currently-open code rows are used. The workbook carries superseded
    1980-1981 rows for the same codes, and reading a closed row as though it
    were current would be a provenance error even where the meaning happens
    to agree.
    """
    import openpyxl  # imported lazily: the workbook is often absent

    workbook_path = Path(path)
    if not workbook_path.exists():
        raise pumd.PumdDataUnavailable(
            f"the PUMD data dictionary is not present: {workbook_path}"
        )
    digest = file_digest(workbook_path)

    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        description = _pubflag_variable_description(workbook)
        meanings, rows = _pubflag_code_meanings(workbook)
    finally:
        workbook.close()

    return PubflagDictionaryReading(
        workbook_sha256=digest,
        workbook_sha256_matches_registry=(digest == pinned_dictionary_digest()),
        sheet=PUBFLAG_DICTIONARY_SHEET,
        variable_description=description,
        code_meanings=meanings,
        rows_read=rows,
    )


def _pubflag_variable_description(workbook) -> str:
    sheet = workbook["Variables"]
    rows = sheet.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    index = {name: i for i, name in enumerate(header)}
    for row in rows:
        if _cell(row, index, "Survey") != PUBFLAG_DICTIONARY_SURVEY:
            continue
        if _cell(row, index, "File") != PUBFLAG_DICTIONARY_FILE:
            continue
        if _cell(row, index, "Variable Name") != PUBFLAG_DICTIONARY_VARIABLE:
            continue
        if row[index["Last year"]] is not None:
            continue  # a superseded definition
        return _cell(row, index, "Variable description")
    raise ShelterSourceError(
        "no open INTERVIEW/MTBI/PUBFLAG row was found on the Variables sheet"
    )


def _pubflag_code_meanings(workbook) -> tuple[dict[str, str], tuple[int, ...]]:
    sheet = workbook[PUBFLAG_DICTIONARY_SHEET]
    rows = sheet.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    index = {name: i for i, name in enumerate(header)}
    meanings: dict[str, str] = {}
    read_at: list[int] = []
    for number, row in enumerate(rows, start=2):
        if _cell(row, index, "Survey") != PUBFLAG_DICTIONARY_SURVEY:
            continue
        if _cell(row, index, "File") != PUBFLAG_DICTIONARY_FILE:
            continue
        if _cell(row, index, "Variable") != PUBFLAG_DICTIONARY_VARIABLE:
            continue
        if row[index["Last year"]] is not None:
            continue  # a superseded code definition
        code = _cell(row, index, "Code value")
        meaning = _cell(row, index, "Code description")
        if code in meanings and meanings[code] != meaning:
            raise ShelterSourceError(
                f"the dictionary gives two open meanings for MTBI PUBFLAG={code}: "
                f"{meanings[code]!r} and {meaning!r}"
            )
        meanings[code] = meaning
        read_at.append(number)
    if not meanings:
        raise ShelterSourceError(
            "no open INTERVIEW/MTBI/PUBFLAG code rows were found on sheet "
            f"{PUBFLAG_DICTIONARY_SHEET!r}"
        )
    return meanings, tuple(read_at)


def _cell(row: Sequence[object], index: Mapping[str, int], name: str) -> str:
    position = index.get(name)
    if position is None or position >= len(row):
        raise ShelterSourceError(f"the dictionary sheet has no column {name!r}")
    value = row[position]
    return "" if value is None else str(value).strip()


# --------------------------------------------------------------------------
# Stage 5: the C1 verdict
# --------------------------------------------------------------------------

#: Verdicts for the phase as a whole.
REPRODUCED = "REPRODUCED"
NOT_REPRODUCED = "NOT_REPRODUCED"


def source_verdict(checks: Sequence[ReproductionCheck]) -> str:
    """REPRODUCED only when every preserved claim was re-derived.

    There is no partial credit and no tolerance. These are exact integer
    counts and single-character codes read from a hash-verified file; a
    near-miss would mean the archive or the reading had changed, which C1
    exists to catch rather than to average away.
    """
    if not checks:
        return NOT_REPRODUCED
    return REPRODUCED if all(c.reproduced for c in checks) else NOT_REPRODUCED
